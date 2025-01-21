import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from .share import tone_and_color_name_to_web_safe_color_code, Rectangle, XlAlignment, Canvas


def edit_canvas(ws, document):
    """キャンバスの編集
    """
    print("🔧　キャンバスの編集")

    # ウィンドウ枠の固定
    ws.freeze_panes = 'C2'

    # Trellis では、タテ：ヨコ＝３：３ で、１ユニットセルとします。
    # また、上辺、右辺、下辺、左辺に、１セル幅の定規を置きます
    canvas_obj = Canvas.from_dict(document['canvas'])
    canvas_rect = canvas_obj.rect_obj

    # 横幅または縦幅が１アウト未満の場合は、定規は描画しません
    if canvas_rect.width_obj.total_of_out_counts_qty < 1 or canvas_rect.height_obj.total_of_out_counts_qty < 1:
        return

    # 行の横幅
    for column_th in range(
            canvas_rect.left_obj.total_of_out_counts_th,
            canvas_rect.left_obj.total_of_out_counts_th + canvas_rect.width_obj.total_of_out_counts_qty):
        column_letter = xl.utils.get_column_letter(column_th)
        ws.column_dimensions[column_letter].width = 2.7    # 2.7 characters = about 30 pixels

    # 列の高さ
    for row_th in range(
            canvas_rect.top_obj.total_of_out_counts_th,
            canvas_rect.top_obj.total_of_out_counts_th + canvas_rect.height_obj.total_of_out_counts_qty):
        ws.row_dimensions[row_th].height = 15    # 15 points = about 30 pixels


def fill_rectangle(ws, column_th, row_th, columns, rows, fill_obj):
    """矩形を塗りつぶします
    """
    # 横へ
    for cur_column_th in range(column_th, column_th + columns):
        column_letter = xl.utils.get_column_letter(cur_column_th)

        # 縦へ
        for cur_row_th in range(row_th, row_th + rows):
            cell = ws[f'{column_letter}{cur_row_th}']
            cell.fill = fill_obj


def draw_xl_border_on_rectangle(ws, xl_border_dict, column_th, row_th, columns, rows):
    """境界線の描画
    """
    top_side = None
    right_side = None
    bottom_side = None
    left_side = None

    # 罫線の style の種類
    # 📖 [openpyxl.styles.borders module](https://openpyxl.readthedocs.io/en/3.1/api/openpyxl.styles.borders.html)
    # ‘mediumDashed’, ‘mediumDashDotDot’, ‘dashDot’, ‘dashed’, ‘slantDashDot’, ‘dashDotDot’, ‘thick’, ‘thin’, ‘dotted’, ‘double’, ‘medium’, ‘hair’, ‘mediumDashDot’

    if 'top' in xl_border_dict and (top_dict := xl_border_dict['top']):
        web_safe_color_code = None
        style = None

        if 'color' in top_dict and (color := top_dict['color']):
            web_safe_color_code = tone_and_color_name_to_web_safe_color_code(color)

        if 'xlStyle' in top_dict and (style := top_dict['xlStyle']):
            pass

        try:
            top_side = Side(style=style, color=ColorSystem.web_safe_color_code_to_xl(web_safe_color_code))
        except:
            print(f'draw_xl_border_on_rectangle: いずれかが、未対応の指定： {style=} {web_safe_color_code=}')


    if 'right' in xl_border_dict and (right_dict := xl_border_dict['right']):
        web_safe_color_code = None
        style = None

        if 'color' in right_dict and (color := right_dict['color']):
            web_safe_color_code = tone_and_color_name_to_web_safe_color_code(color)

        if 'xlStyle' in right_dict and (style := right_dict['xlStyle']):
            pass

        try:
            right_side = Side(style=style, color=ColorSystem.web_safe_color_code_to_xl(web_safe_color_code))
        except:
            print(f'draw_xl_border_on_rectangle: いずれかが、未対応の指定： {style=} {web_safe_color_code=}')


    if 'bottom' in xl_border_dict and (bottom_dict := xl_border_dict['bottom']):
        web_safe_color_code = None
        style = None

        if 'color' in bottom_dict and (color := bottom_dict['color']):
            web_safe_color_code = tone_and_color_name_to_web_safe_color_code(color)

        if 'xlStyle' in bottom_dict and (style := bottom_dict['xlStyle']):
            pass

        try:
            bottom_side = Side(style=style, color=ColorSystem.web_safe_color_code_to_xl(web_safe_color_code))
        except:
            print(f'draw_xl_border_on_rectangle: いずれかが、未対応の指定： {style=} {web_safe_color_code=}')


    if 'left' in xl_border_dict and (left_dict := xl_border_dict['left']):
        web_safe_color_code = None
        style = None

        if 'color' in left_dict and (color := left_dict['color']):
            web_safe_color_code = tone_and_color_name_to_web_safe_color_code(color)

        if 'xlStyle' in left_dict and (style := left_dict['xlStyle']):
            pass

        try:
            left_side = Side(style=style, color=ColorSystem.web_safe_color_code_to_xl(web_safe_color_code))
        except:
            print(f'draw_xl_border_on_rectangle: いずれかが、未対応の指定： {style=} {web_safe_color_code=}')


    # TODO 厚みが１のケースや、角は、２辺に線を引く

    
    top_border = Border(top=top_side)           # 上辺
    right_border = Border(right=right_side)     # 右辺
    bottom_border = Border(bottom=bottom_side)  # 下辺
    left_border = Border(left=left_side)        # 左辺

    # 水平方向
    if rows == 0 or rows == 1:
        if rows == 0:
            # 上辺だけ引く
            horizontal_border = Border(top=top_side)
        else:
            # 上辺と下辺の両方を引く
            horizontal_border = Border(top=top_side, bottom=bottom_side)

        # （角を除く）横へ
        for cur_column_th in range(column_th + 1, column_th + columns - 1):
            column_letter = xl.utils.get_column_letter(cur_column_th)
            cell = ws[f'{column_letter}{row_th}']
            cell.border = horizontal_border

    # 上辺を引くのと、下辺を引くのとがある
    else:
        top_border = Border(top=top_side)
        bottom_border = Border(bottom=bottom_side)

        # （角を除く）横へ
        for cur_column_th in range(column_th + 1, column_th + columns - 1):
            column_letter = xl.utils.get_column_letter(cur_column_th)

            cell = ws[f'{column_letter}{row_th}']
            cell.border = top_border

            cell = ws[f'{column_letter}{row_th + rows - 1}']
            cell.border = bottom_border


    # 垂直方向
    if columns == 0 or columns == 1:
        if columns == 0:
            # 左辺だけ引く
            vertical_border = Border(left=left_side)
        else:
            # 左辺と右辺の両方を引く
            vertical_border = Border(left=left_side, right=right_side)

        # （角を除く）縦へ
        for cur_row_th in range(row_th + 1, row_th + rows - 1):
            column_letter = xl.utils.get_column_letter(columns)
            cell = ws[f'{column_letter}{cur_row_th}']
            cell.border = vertical_border

    # 左辺を引くのと、右辺を引くのとがある
    else:
        left_border = Border(left=left_side)
        right_border = Border(right=right_side)

        # （角を除く）縦へ
        for cur_row_th in range(row_th + 1, row_th + rows - 1):
            column_letter = xl.utils.get_column_letter(column_th)
            cell = ws[f'{column_letter}{cur_row_th}']
            cell.border = left_border

            column_letter = xl.utils.get_column_letter(column_th + columns - 1)
            cell = ws[f'{column_letter}{cur_row_th}']
            cell.border = right_border


    # 左上隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th}']
        cell.border = Border(top=top_side, left=left_side)

    # 右上隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th + columns - 1)
        cell = ws[f'{column_letter}{row_th}']
        cell.border = Border(top=top_side, right=right_side)

    # 左下隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th + rows - 1}']
        cell.border = Border(left=left_side, bottom=bottom_side)

    # 右下隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th + columns - 1)
        cell = ws[f'{column_letter}{row_th + rows - 1}']
        cell.border = Border(right=right_side, bottom=bottom_side)

    # 四方
    if columns == 1 and rows == 1:
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th}']
        cell.border = Border(top=top_side, right=right_side, bottom=bottom_side, left=left_side)


def print_text(ws, location_obj, text, xl_alignment_obj, xl_font_obj):
    """テキスト描画
    """

    # テキストの位置
    column_th = location_obj.x_obj.total_of_out_counts_th
    row_th = location_obj.y_obj.total_of_out_counts_th

    # テキスト設定
    column_letter = xl.utils.get_column_letter(column_th)
    cell = ws[f'{column_letter}{row_th}']
    cell.value = text

    # フォント設定
    if xl_font_obj:
        cell.font = Font(color=xl_font_obj.color_code_for_xl)

    # テキストの位置揃え
    if xl_alignment_obj:
        cell.alignment = Alignment(
                horizontal=xl_alignment_obj.xlHorizontal,
                vertical=xl_alignment_obj.xlVertical)
