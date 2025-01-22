import os
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as XlImage
import json


class Share():
    """共有
    """

    @classmethod
    @property
    def OUT_COUNTS_THAT_CHANGE_INNING(clazz):
        """3 ということが言いたいだけの、長い定数名。
        Trellis では、3x3cells で［大グリッド１マス分］とします
        """
        return 3


######################
# MARK: InningsPitched
######################
class InningsPitched():
    """投球回。
    トレリスでは、セル番号を指定するのに使っている
    """


    @staticmethod
    def from_integer_and_decimal_part(integer_part, decimal_part):
        """整数部と小数部を指定
        """
        return InningsPitched(integer_part=integer_part, decimal_part=decimal_part)


    @staticmethod
    def from_var_value(var_value):

        try:
            # "100" が来たら 100 にする
            var_value = int(var_value)
        except ValueError:
            pass

        if isinstance(var_value, int):
            return InningsPitched(
                    integer_part=var_value,
                    decimal_part=0)

        elif isinstance(var_value, str):
            integer_part, decimal_part = map(int, var_value.split('o', 2))
            return InningsPitched(
                    integer_part=integer_part,
                    decimal_part=decimal_part)

        else:
            raise ValueError(f'{type(var_value)=} {var_value=}')

        return InningsPitched(var_value)


    def __init__(self, integer_part, decimal_part):
        self._integer_part = integer_part
        self._decimal_part = decimal_part

        if self._decimal_part == 0:
            self._var_value = self._integer_part
        else:
            self._var_value = f'{self._integer_part}o{self._decimal_part}'

        self._total_of_out_counts_qty = self._integer_part * Share.OUT_COUNTS_THAT_CHANGE_INNING + self._decimal_part


    @property
    def var_value(self):
        """投球回の整数だったり、"3o2" 形式の文字列だったりします
        """
        return self._var_value


    @property
    def integer_part(self):
        """投球回の整数部"""
        return self._integer_part


    @property
    def decimal_part(self):
        """投球回の小数部"""
        return self._decimal_part


    @property
    def total_of_out_counts_qty(self):
        """0から始まるアウト・カウントの総数
        """
        return self._total_of_out_counts_qty


    @property
    def total_of_out_counts_th(self):
        """1から始まるアウト・カウントの総数
        """
        return self._total_of_out_counts_qty + 1


    def offset(self, var_value):
        """この投球回に、引数を加算した数を算出して返します"""
        l = self                       # Left operand
        r = InningsPitched.from_var_value(var_value)  # Right operand
        sum_decimal_part = l.decimal_part + r.decimal_part
        integer_part = l.integer_part + r.integer_part + sum_decimal_part // Share.OUT_COUNTS_THAT_CHANGE_INNING
        return InningsPitched.from_integer_and_decimal_part(
                integer_part=integer_part,
                decimal_part=sum_decimal_part % Share.OUT_COUNTS_THAT_CHANGE_INNING)


#############
# MARK: Point
#############
class Point():
    """位置
    """


    @staticmethod
    def from_dict(point_dict):
        """辞書を元に生成
        """

        try:
            first_x = point_dict['x']
        except:
            print(f'ERROR: Rectangle.from_dict: {point_dict=}')
            raise

        second_x = 0
        if isinstance(first_x, str):
            first_x, second_x = map(int, first_x.split('o', 2))

        first_y = point_dict['y']
        second_y = 0
        if isinstance(first_y, str):
            first_y, second_y = map(int, first_y.split('o', 2))

        return Point(
                first_x=first_x,
                second_x=second_x,
                first_y=first_y,
                second_y=second_y)


    def __init__(self, first_x, second_x, first_y, second_y):
        """初期化
        """
        self._x_obj = InningsPitched.from_integer_and_decimal_part(integer_part=first_x, decimal_part=second_x)
        self._y_obj = InningsPitched.from_integer_and_decimal_part(integer_part=first_y, decimal_part=second_y)


    @property
    def x_obj(self):
        return self._x_obj


    @property
    def y_obj(self):
        return self._y_obj


#################
# MARK: Rectangle
#################
class Rectangle():
    """矩形
    """


    @staticmethod
    def from_dict(bounds_dict):
        """ラインテープのセグメントの矩形情報を取得
        """

        try:
            main_left = bounds_dict['left']
        except:
            print(f'ERROR: Rectangle.from_dict: {bounds_dict=}')
            raise

        sub_left = 0
        if isinstance(main_left, str):
            main_left, sub_left = map(int, main_left.split('o', 2))

        main_top = bounds_dict['top']
        sub_top = 0
        if isinstance(main_top, str):
            main_top, sub_top = map(int, main_top.split('o', 2))

        # right は、その数を含まない。
        # right が指定されていれば、 width より優先する
        if 'right' in bounds_dict:
            right = bounds_dict['right']
            sub_right = 0
            if isinstance(right, str):
                right, sub_right = map(int, right.split('o', 2))

            main_width = right - main_left
            sub_width = sub_right - sub_left

        else:
            main_width = bounds_dict['width']
            sub_width = 0
            if isinstance(main_width, str):
                main_width, sub_width = map(int, main_width.split('o', 2))

        # bottom は、その数を含まない。
        # bottom が指定されていれば、 width より優先する
        if 'bottom' in bounds_dict:
            bottom = bounds_dict['bottom']
            sub_bottom = 0
            if isinstance(bottom, str):
                bottom, sub_bottom = map(int, bottom.split('o', 2))

            main_height = bottom - main_top
            sub_height = sub_bottom - sub_top

        else:
            main_height = bounds_dict['height']
            sub_height = 0
            if isinstance(main_height, str):
                main_height, sub_height = map(int, main_height.split('o', 2))

        return Rectangle(
                main_left=main_left,
                sub_left=sub_left,
                main_top=main_top,
                sub_top=sub_top,
                main_width=main_width,
                sub_width=sub_width,
                main_height=main_height,
                sub_height=sub_height)


    def __init__(self, main_left, sub_left, main_top, sub_top, main_width, sub_width, main_height, sub_height):
        """初期化
        """
        self._left_obj = InningsPitched.from_integer_and_decimal_part(integer_part=main_left, decimal_part=sub_left)
        self._top_obj = InningsPitched.from_integer_and_decimal_part(integer_part=main_top, decimal_part=sub_top)
        self._width_obj = InningsPitched.from_integer_and_decimal_part(integer_part=main_width, decimal_part=sub_width)
        self._height_obj = InningsPitched.from_integer_and_decimal_part(integer_part=main_height, decimal_part=sub_height)
        self._right_obj = None
        self._bottom_obj = None


    def _calculate_right(self):
        sum_decimal_part = self._left_obj.decimal_part + self._width_obj.decimal_part
        self._right_obj = InningsPitched.from_integer_and_decimal_part(
                integer_part=self._left_obj.integer_part + self._width_obj.integer_part + sum_decimal_part // Share.OUT_COUNTS_THAT_CHANGE_INNING,
                decimal_part=sum_decimal_part % Share.OUT_COUNTS_THAT_CHANGE_INNING)


    def _calculate_bottom(self):
        sum_decimal_part = self._top_obj.decimal_part + self._height_obj.decimal_part
        self._bottom_obj = InningsPitched.from_integer_and_decimal_part(
                integer_part=self._top_obj.integer_part + self._height_obj.integer_part + sum_decimal_part // Share.OUT_COUNTS_THAT_CHANGE_INNING,
                decimal_part=sum_decimal_part % Share.OUT_COUNTS_THAT_CHANGE_INNING)


    @property
    def left_obj(self):
        return self._left_obj


    @property
    def right_obj(self):
        """矩形の右位置
        """
        if not self._right_obj:
            self._calculate_right()
        return self._right_obj


    @property
    def top_obj(self):
        return self._top_obj


    @property
    def bottom_obj(self):
        """矩形の下位置
        """
        if not self._bottom_obj:
            self._calculate_bottom()
        return self._bottom_obj


    @property
    def width_obj(self):
        return self._width_obj


    @property
    def height_obj(self):
        return self._height_obj


####################
# MARK: Color system
####################
class ColorSystem():
    """色システム
    """


    _none_pattern_fill = PatternFill(patternType=None)

    @classmethod
    @property
    def none_pattern_fill(clazz):
        return clazz._none_pattern_fill


    # エクセルの色システム（勝手に作ったったもの）
    _xl_color_code_to_web_safe_color_dict = {
        'xlTheme' : {
            'xlWhite' : '#FFFFFF',
            'xlBlack' : '#000000',
            'xlRed_gray' : '#E7E6E6',
            'xlBlueGray' : '#44546A',
            'xlBlue' : '#5B9BD5',
            'xlRed' : '#ED7D31',
            'xlGray' : '#A5A5A5',
            'xlYellow' : '#FFC000',
            'xlNaviy' : '#4472C4',
            'xlGreen' : '#70AD47',
        },
        'xlPale' : {
            'xlWhite' : '#F2F2F2',
            'xlBlack' : '#808080',
            'xlRed_gray' : '#AEAAAA',
            'xlBlueGray' : '#D6DCE4',
            'xlBlue' : '#DDEBF7',
            'xlRed' : '#FCE4D6',
            'xlGray' : '#EDEDED',
            'xlYellow' : '#FFF2CC',
            'xlNaviy' : '#D9E1F2',
            'xlGreen' : '#E2EFDA',
        },
        'xlLight' : {
            'xlWhite' : '#D9D9D9',
            'xlBlack' : '#595959',
            'xlRed_gray' : '#757171',
            'xlBlueGray' : '#ACB9CA',
            'xlBlue' : '#BDD7EE',
            'xlRed' : '#F8CBAD',
            'xlGray' : '#DBDBDB',
            'xlYellow' : '#FFE699',
            'xlNaviy' : '#B4C6E7',
            'xlGreen' : '#C6E0B4',
        },
        'xlSoft' : {
            'xlWhite' : '#BFBFBF',
            'xlBlack' : '#404040',
            'xlRed_gray' : '#3A3838',
            'xlBlueGray' : '#8497B0',
            'xlBlue' : '#9BC2E6',
            'xlRed' : '#F4B084',
            'xlGray' : '#C9C9C9',
            'xlYellow' : '#FFD966',
            'xlNaviy' : '#8EA9DB',
            'xlGreen' : '#A9D08E',
        },
        'xlStrong' : {
            'xlWhite' : '#A6A6A6',
            'xlBlack' : '#262626',
            'xlRed_gray' : '#3A3838',
            'xlBlueGray' : '#333F4F',
            'xlBlue' : '#2F75B5',
            'xlRed' : '#C65911',
            'xlGray' : '#7B7B7B',
            'xlYellow' : '#BF8F00',
            'xlNaviy' : '#305496',
            'xlGreen' : '#548235',
        },
        'xlDeep' : {
            'xlWhite' : '#808080',
            'xlBlack' : '#0D0D0D',
            'xlRed_gray' : '#161616',
            'xlBlueGray' : '#161616',
            'xlBlue' : '#1F4E78',
            'xlRed' : '#833C0C',
            'xlGray' : '#525252',
            'xlYellow' : '#806000',
            'xlNaviy' : '#203764',
            'xlGreen' : '#375623',
        },
        'xlStandard' : {
            'xlRed' : '#C00000',
            'xlRed' : '#FF0000',
            'xlOrange' : '#FFC000',
            'xlYellow' : '#FFFF00',
            'xlYellowGreen' : '#92D050',
            'xlGreen' : '#00B050',
            'xlDodgerBlue' : '#00B0F0',
            'xlBlue' : '#0070C0',
            'xlNaviy' : '#002060',
            'xlViolet' : '#7030A0',
        }
    }

    @classmethod
    @property
    def xl_color_code_to_web_safe_color_dict(clazz):
        return clazz._xl_color_code_to_web_safe_color_dict


    @staticmethod
    def web_safe_color_code_to_xl(web_safe_color_code):
        """頭の `#` を外します
        """
        return web_safe_color_code[1:]


    @staticmethod
    @property
    def AUTO():
        return 'auto'


    @staticmethod
    @property
    def DARKNESS():
        return 'darkness'


    @staticmethod
    @property
    def PAPER_COLOR():
        return 'paperColor'


    @staticmethod
    @property
    def TONE_AND_COLOR_NAME():
        return 'toneAndColorName'


    @staticmethod
    @property
    def WEB_SAFE_COLOR():
        return 'webSafeColor'


    @staticmethod
    def what_is_var_color_name(var_color_name):
        """TODO トーン名・色名の欄に何が入っているか判定します
        """

        # 何も入っていない、または False が入っている
        if not var_color_name:
            return False

        # ナンが入っている
        if var_color_name is None:
            return None

        # ウェブ・セーフ・カラーが入っている
        #
        #   とりあえず、 `#` で始まるなら、ウェブセーフカラーとして扱う
        #
        if var_color_name.startswith('#'):
            return ColorSystem.WEB_SAFE_COLOR

        # 色相名と色名だ
        if '.' in var_color_name:
            return ColorSystem.TONE_AND_COLOR_NAME

        # "auto", "paperColor" キーワードのいずれかが入っている
        if var_color_name in [ColorSystem.AUTO, ColorSystem.PAPER_COLOR]:
            return var_color_name
        
        raise ValueError(f"""ERROR: what_is_var_color_name: undefined {var_color_name=}""")


    @staticmethod
    def solve_tone_and_color_name(tone_and_color_name):
        try:
            tone, color = tone_and_color_name.split('.', 2)
        except:
            print(f'solve_tone_and_color_name: tone.color の形式でない {tone_and_color_name=}')
            raise


        tone = tone.strip()
        color = color.strip()

        if tone in ColorSystem.xl_color_code_to_web_safe_color_dict and (tone_dict := ColorSystem.xl_color_code_to_web_safe_color_dict[tone]):
            if color in tone_dict and (web_safe_color_code := tone_dict[color]):
                return web_safe_color_code

        print(f'var_color_name_to_web_safe_color_code: 色がない {tone_and_color_name=}')
        return None



    @staticmethod
    def var_color_name_to_web_safe_color_code(var_color_name):
        """様々な色名をウェブ・セーフ・カラーの１６進文字列の色コードに変換します
        """

        # 色が指定されていないとき、この関数を呼び出してはいけません
        if var_color_name is None:
            raise Exception(f'var_color_name_to_web_safe_color_code: 色が指定されていません')

        # 背景色を［なし］にします。透明（transparent）で上書きするのと同じです
        if var_color_name == 'paperColor':
            raise Exception(f'var_color_name_to_web_safe_color_code: 透明色には対応していません')

        # ［auto］は自動で影の色を設定する機能ですが、その機能をオフにしているときは、とりあえず黒色にします
        if var_color_name == 'auto':
            return ColorSystem.xl_color_code_to_web_safe_color_dict['xlTheme']['xlBlack']

        # `#` で始まるなら、ウェブセーフカラーとして扱う
        if var_color_name.startswith('#'):
            return var_color_name


        return ColorSystem.solve_tone_and_color_name(tone_and_color_name=var_color_name)


    @staticmethod
    def var_color_name_to_fill_obj(var_color_name):
        """様々な色名を FillPattern オブジェクトに変換します
        """

        # 色が指定されていないとき、この関数を呼び出してはいけません
        if var_color_name is None:
            raise Exception(f'var_color_name_to_fill_obj: 色が指定されていません')

        # 背景色を［なし］にします。透明（transparent）で上書きするのと同じです
        if var_color_name == 'paperColor':
            return ColorSystem.none_pattern_fill

        # ［auto］は自動で影の色を設定する機能ですが、その機能をオフにしているときは、とりあえず黒色にします
        if var_color_name == 'auto':
            return PatternFill(
                    patternType='solid',
                    fgColor=ColorSystem.web_safe_color_code_to_xl(ColorSystem.xl_color_code_to_web_safe_color_dict['xlTheme']['xlBlack']))

        try:
            tone, color = var_color_name.split('.', 2)
        except:
            print(f'ERROR: {var_color_name=}')
            raise

        tone = tone.strip()
        color = color.strip()

        if tone in ColorSystem.xl_color_code_to_web_safe_color_dict:
            if color in ColorSystem.xl_color_code_to_web_safe_color_dict[tone]:
                return PatternFill(
                        patternType='solid',
                        fgColor=ColorSystem.web_safe_color_code_to_xl(ColorSystem.xl_color_code_to_web_safe_color_dict[tone][color]))

        print(f'var_color_name_to_fill_obj: 色がない {var_color_name=}')
        return ColorSystem.none_pattern_fill


    _darkness_dict = {        
    }


    @classmethod
    def set_color_system(clazz, ws, document):
        """TODO 色システムの設定
        """

        if 'colorSystem' in document and (color_system_dict := document['colorSystem']):
            if 'darkness' in color_system_dict and (darkness_dict := color_system_dict['darkness']):
                darkness_dict_edit = dict(darkness_dict)

                # TODO 変換できる色名は、変換したい
                for var_color_name_before_change, var_color_name_after_change in darkness_dict_edit.items():

                    color_type = ColorSystem.what_is_var_color_name(var_color_name_before_change)
                    if color_type == ColorSystem.TONE_AND_COLOR_NAME:
                        tlanslated_var_color_name_before_change = ColorSystem.solve_tone_and_color_name(tone_and_color_name=var_color_name_before_change)

                        if var_color_name_before_change != tlanslated_var_color_name_before_change:
                            old_value = darkness_dict_edit[var_color_name_before_change]
                            del darkness_dict_edit[var_color_name_before_change]
                            darkness_dict_edit[tlanslated_var_color_name_before_change] = old_value

                clazz._darkness_dict = darkness_dict_edit


    @staticmethod
    def solve_darkness(darkness, web_safe_color_code):
        """TODO
        """
        return


###################
# MARK: XlAlignment
###################
class XlAlignment():
    """Excel 用テキストの位置揃え
    """


    @staticmethod
    def from_dict(xl_alignment_dict):
        """辞書を元に生成

        📖 [openpyxl.styles.alignment module](https://openpyxl.readthedocs.io/en/latest/api/openpyxl.styles.alignment.html)
        horizontal: Value must be one of {‘fill’, ‘left’, ‘distributed’, ‘justify’, ‘center’, ‘general’, ‘centerContinuous’, ‘right’}
        vertical: Value must be one of {‘distributed’, ‘justify’, ‘center’, ‘bottom’, ‘top’}
        """
        xlHorizontal = None
        xlVertical = None
        if 'xlHorizontal' in xl_alignment_dict:
            xlHorizontal = xl_alignment_dict['xlHorizontal']

        if 'xlVertical' in xl_alignment_dict:
            xlVertical = xl_alignment_dict['xlVertical']

        return XlAlignment(
                xlHorizontal=xlHorizontal,
                xlVertical=xlVertical)


    def __init__(self, xlHorizontal, xlVertical):
        self._xl_horizontal = xlHorizontal
        self._xl_vertical = xlVertical


    @property
    def xlHorizontal(self):
        return self._xl_horizontal


    @property
    def xlVertical(self):
        return self._xl_vertical


##############
# MARK: XlFont
##############
class XlFont():
    """Excel 用フォント
    """


    @staticmethod
    def from_dict(xl_font_dict):
        """辞書を元に生成
        """
        web_safe_color_code = None
        if 'color' in xl_font_dict:
            web_safe_color_code = ColorSystem.var_color_name_to_web_safe_color_code(xl_font_dict['color'])

        return XlFont(
                web_safe_color_code=web_safe_color_code)


    def __init__(self, web_safe_color_code):
        self._web_safe_color_code = web_safe_color_code


    @property
    def web_safe_color_code(self):
        return self._web_safe_color_code


    @property
    def color_code_for_xl(self):
        return ColorSystem.web_safe_color_code_to_xl(self._web_safe_color_code)


##############
# MARK: Canvas
##############
class Canvas():
    """キャンバス
    """


    def from_dict(canvas_dict):

        rect_obj = None
        if 'bounds' in canvas_dict and (bounds_dict := canvas_dict['bounds']):
            rect_obj = Rectangle.from_dict(bounds_dict)

        return Canvas(
                rect_obj=rect_obj)


    def __init__(self, rect_obj):
        self._rect_obj = rect_obj


    @property
    def rect_obj(self):
        return self._rect_obj


##############
# MARK: Pillar
##############
class Pillar():
    """柱
    """


    def from_dict(pillar_dict):

        rect_obj = None
        if 'bounds' in pillar_dict and (bounds_dict := pillar_dict['bounds']):
            rect_obj = Rectangle.from_dict(bounds_dict)

        # FIXME: if 'baseColor' in pillar_dict and (var_color_name := pillar_dict['baseColor']):


        return Canvas(
                rect_obj=rect_obj)


    def __init__(self, rect_obj):
        self._rect_obj = rect_obj


    @property
    def rect_obj(self):
        return self._rect_obj


############
# MARK: Card
############
class Card():
    """カード
    """


    def from_dict(card_dict):

        rect_obj = None
        if 'bounds' in card_dict and (bounds_dict := card_dict['bounds']):
            rect_obj = Rectangle.from_dict(bounds_dict)

        return Canvas(
                rect_obj=rect_obj)


    def __init__(self, rect_obj):
        self._rect_obj = rect_obj


    @property
    def rect_obj(self):
        return self._rect_obj


################
# MARK: Terminal
################
class Terminal():
    """端子
    """


    def from_dict(terminal_dict):

        rect_obj = None
        if 'bounds' in terminal_dict and (bounds_dict := terminal_dict['bounds']):
            rect_obj = Rectangle.from_dict(bounds_dict)

        return Canvas(
                rect_obj=rect_obj)


    def __init__(self, rect_obj):
        self._rect_obj = rect_obj


    @property
    def rect_obj(self):
        return self._rect_obj
