import re

from openpyxl.styles import PatternFill


class ColorSystem():
    """色システム
    """


    _none_pattern_fill = PatternFill(patternType=None)

    @classmethod
    @property
    def none_pattern_fill(clazz):
        return clazz._none_pattern_fill


    @classmethod
    def alias_to_web_safe_color_dict(clazz, contents_doc):
        # # TODO 前もって作っておきたい
        # if 'colorSystem' not in contents_doc:
        #     return {}
        
        # if 'alias' not in contents_doc['colorSystem']:
        #     return {}

        return contents_doc['colorSystem']['alias']


    @staticmethod
    def web_safe_color_code_to_xl(web_safe_color_code):
        """頭の `#` を外します
        """

        # FIXME チェック★
        if not re.match(r'^#[0-9a-fA-F]{6}$', web_safe_color_code):
            raise ValueError(f'web_safe_color_code_to_xl: ウェブ・セーフ・カラーじゃないかも？ {web_safe_color_code=}')
        
        #print(f'★ {web_safe_color_code=}')

        return web_safe_color_code[1:]


    @classmethod
    @property
    def AUTO(clazz):
        return 'auto'


    @classmethod
    @property
    def DARKNESS(clazz):
        return 'darkness'


    @classmethod
    @property
    def PAPER_COLOR(clazz):
        return 'paperColor'


    @classmethod
    @property
    def TONE_AND_COLOR_NAME(clazz):
        return 'toneAndColorName'


    @classmethod
    @property
    def WEB_SAFE_COLOR_CODE(clazz):
        return 'webSafeColorCode'


    @classmethod
    @property
    def XL_COLOR_CODE(clazz):
        return 'xlColorCode'


    @staticmethod
    def what_is_var_color_name(var_color_name):
        """TODO トーン名・色名の欄に何が入っているか判定します
        """

        # 何も入っていない、または False が入っている
        if not var_color_name:
            return False

        # ナンが入っている
        if var_color_name is None:
            return None

        # ウェブ・セーフ・カラーが入っている
        #
        #   とりあえず、 `#` で始まるなら、ウェブセーフカラーとして扱う
        #
        #if var_color_name.startswith('#'):
        if re.match(r'^#[0-9a-fA-f]{6}$', var_color_name):
            return ColorSystem.WEB_SAFE_COLOR_CODE

        if re.match(r'^[0-9a-fA-f]{6}$', var_color_name):
            return ColorSystem.XL_COLOR_CODE

        # 色相名と色名だ
        #if '.' in var_color_name:
        if re.match(r'^[0-9a-zA-Z_]+\.[0-9a-zA-Z_]+$', var_color_name):
            return ColorSystem.TONE_AND_COLOR_NAME

        # "auto", "paperColor" キーワードのいずれかが入っている
        if var_color_name in [ColorSystem.AUTO, ColorSystem.PAPER_COLOR]:
            return var_color_name
        
        raise ValueError(f"""ERROR: what_is_var_color_name: undefined {var_color_name=} {ColorSystem.AUTO=} {ColorSystem.PAPER_COLOR=}""")


    @staticmethod
    def solve_tone_and_color_name(contents_doc, tone_and_color_name):
        try:
            tone, color = tone_and_color_name.split('.', 2)
        except:
            print(f'solve_tone_and_color_name: tone.color の形式でない {tone_and_color_name=}')
            raise


        tone = tone.strip()
        color = color.strip()

        if tone in ColorSystem.alias_to_web_safe_color_dict(contents_doc) and (tone_dict := ColorSystem.alias_to_web_safe_color_dict(contents_doc)[tone]):
            if color in tone_dict and (web_safe_color_code := tone_dict[color]):
                return web_safe_color_code

        print(f'var_color_name_to_web_safe_color_code: 色がない {tone_and_color_name=}')
        return None



    @staticmethod
    def var_color_name_to_web_safe_color_code(contents_doc, var_color_name):
        """様々な色名をウェブ・セーフ・カラーの１６進文字列の色コードに変換します
        """

        color_type = ColorSystem.what_is_var_color_name(var_color_name=var_color_name)

        # 色が指定されていないとき、この関数を呼び出してはいけません
        if not color_type:
            raise Exception(f'var_color_name_to_web_safe_color_code: 色が指定されていません')

        # 背景色を［なし］にします。透明（transparent）で上書きするのと同じです
        if color_type == ColorSystem.PAPER_COLOR:
            raise Exception(f'var_color_name_to_web_safe_color_code: 透明色には対応していません')

        # ［auto］は自動で影の色を設定する機能ですが、その機能をオフにしているときは、とりあえず黒色にします
        if color_type == ColorSystem.AUTO:
            return ColorSystem.alias_to_web_safe_color_dict(contents_doc=contents_doc)['xlTheme']['xlBlack']

        # ウェブセーフカラー
        if color_type == ColorSystem.WEB_SAFE_COLOR_CODE:
            return var_color_name

        return ColorSystem.solve_tone_and_color_name(
            contents_doc=contents_doc,
            tone_and_color_name=var_color_name)


    @staticmethod
    def var_color_name_to_fill_obj(contents_doc, var_color_name):
        """様々な色名を FillPattern オブジェクトに変換します
        """

        color_type = ColorSystem.what_is_var_color_name(var_color_name=var_color_name)

        # 色が指定されていないとき、この関数を呼び出してはいけません
        if not color_type:
            raise Exception(f'var_color_name_to_fill_obj: 色が指定されていません')

        # 背景色を［なし］にします。透明（transparent）で上書きするのと同じです
        if color_type == ColorSystem.PAPER_COLOR:
            return ColorSystem.none_pattern_fill

        if color_type == ColorSystem.XL_COLOR_CODE:
            return PatternFill(
                    patternType='solid',
                    fgColor=var_color_name)

        # ［auto］は自動で影の色を設定する機能ですが、その機能をオフにしているときは、とりあえず黒色にします
        if color_type == ColorSystem.AUTO:
            xl_color_name = ColorSystem.web_safe_color_code_to_xl(
                    ColorSystem.alias_to_web_safe_color_dict(contents_doc)['xlTheme']['xlBlack'])

            #if not re.match(r'^[0-9a-fA-f]{6}$', xl_color_name): #FIXME
            #    raise ValueError(f'色指定がおかしい {xl_color_name=}')
            # else:
            #     print(f'★ {xl_color_name=}')

            return PatternFill(
                    patternType='solid',
                    fgColor=xl_color_name)

        # ウェブ・セーフ・カラーを、エクセルの引数書式へ変換
        if color_type == ColorSystem.WEB_SAFE_COLOR_CODE:
            return PatternFill(
                    patternType='solid',
                    fgColor=ColorSystem.web_safe_color_code_to_xl(var_color_name))

        if color_type == ColorSystem.TONE_AND_COLOR_NAME:
            tone, color = var_color_name.split('.', 2)
            tone = tone.strip()
            color = color.strip()

            if tone in ColorSystem.alias_to_web_safe_color_dict(contents_doc):
                if color in ColorSystem.alias_to_web_safe_color_dict(contents_doc)[tone]:
                    return PatternFill(
                            patternType='solid',
                            fgColor=ColorSystem.web_safe_color_code_to_xl(ColorSystem.alias_to_web_safe_color_dict(contents_doc)[tone][color]))


        print(f'var_color_name_to_fill_obj: 色がない {var_color_name=}')
        return ColorSystem.none_pattern_fill
